import openpyxl
import pytest
from openpyxl.utils import get_column_letter

from xlsxtpl.col_expander import ColExpander


@pytest.fixture
def ws_with_body():
    """Workbook with a 3-column body (cols B-D) surrounded by a header and footer col."""
    wb = openpyxl.Workbook()
    ws = wb.active
    # Col A (1) = header
    ws["A1"] = "Header"
    # Body: cols B(2), C(3), D(4) across 2 rows
    ws["B1"] = "Col B R1"
    ws["B2"] = 10
    ws["C1"] = "Col C R1"
    ws["C2"] = 20
    ws["D1"] = "Col D R1"
    ws["D2"] = 30
    # Col E (5) = footer
    ws["E1"] = "Footer"
    # Set a custom width on column B
    ws.column_dimensions["B"].width = 25
    return ws


class TestExpandForLoop:
    def test_single_iteration_no_change(self, ws_with_body):
        expander = ColExpander(ws_with_body)
        added = expander.expand_for_loop(body_start=2, body_end=4, iteration_count=1)
        assert added == 0
        assert ws_with_body.max_column == 5

    def test_zero_iterations_no_change(self, ws_with_body):
        expander = ColExpander(ws_with_body)
        added = expander.expand_for_loop(body_start=2, body_end=4, iteration_count=0)
        assert added == 0

    def test_two_iterations(self, ws_with_body):
        ws = ws_with_body
        expander = ColExpander(ws)
        added = expander.expand_for_loop(body_start=2, body_end=4, iteration_count=2)
        assert added == 3  # (2-1) * 3 = 3

        # Original body preserved in cols B(2)-D(4)
        assert ws.cell(row=1, column=2).value == "Col B R1"
        assert ws.cell(row=1, column=3).value == "Col C R1"
        assert ws.cell(row=1, column=4).value == "Col D R1"

        # Copy in cols E(5)-G(7)
        assert ws.cell(row=1, column=5).value == "Col B R1"
        assert ws.cell(row=2, column=5).value == 10
        assert ws.cell(row=1, column=6).value == "Col C R1"
        assert ws.cell(row=1, column=7).value == "Col D R1"

        # Footer shifted to col H(8)
        assert ws.cell(row=1, column=8).value == "Footer"

    def test_three_iterations(self, ws_with_body):
        ws = ws_with_body
        expander = ColExpander(ws)
        added = expander.expand_for_loop(body_start=2, body_end=4, iteration_count=3)
        assert added == 6  # (3-1) * 3 = 6

        # Third iteration in cols H(8)-J(10)
        assert ws.cell(row=1, column=8).value == "Col B R1"
        assert ws.cell(row=1, column=9).value == "Col C R1"
        assert ws.cell(row=1, column=10).value == "Col D R1"

        # Footer at col K(11)
        assert ws.cell(row=1, column=11).value == "Footer"

    def test_style_preservation(self, ws_with_body):
        ws = ws_with_body
        ws.cell(row=1, column=2).font = openpyxl.styles.Font(bold=True)
        expander = ColExpander(ws)
        expander.expand_for_loop(body_start=2, body_end=4, iteration_count=2)
        # Copied column should have bold font
        assert ws.cell(row=1, column=5).font.bold is True

    def test_column_width_preservation(self, ws_with_body):
        ws = ws_with_body
        expander = ColExpander(ws)
        expander.expand_for_loop(body_start=2, body_end=4, iteration_count=2)
        # Col E(5) (first col of iteration 2) should have same width as col B(2)
        assert ws.column_dimensions[get_column_letter(5)].width == 25

    def test_multi_row_body(self):
        """Body spans multiple rows — all rows get duplicated."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = "R1C1"
        ws.cell(row=2, column=1).value = "R2C1"
        ws.cell(row=1, column=2).value = "Footer"

        expander = ColExpander(ws)
        added = expander.expand_for_loop(body_start=1, body_end=1, iteration_count=3)
        assert added == 2

        # Three copies of column 1
        assert ws.cell(row=1, column=1).value == "R1C1"
        assert ws.cell(row=1, column=2).value == "R1C1"
        assert ws.cell(row=1, column=3).value == "R1C1"
        assert ws.cell(row=2, column=1).value == "R2C1"
        assert ws.cell(row=2, column=2).value == "R2C1"
        assert ws.cell(row=2, column=3).value == "R2C1"
        # Footer shifted
        assert ws.cell(row=1, column=4).value == "Footer"


class TestRemoveCols:
    def test_remove_single_col(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = "keep"
        ws.cell(row=1, column=2).value = "delete"
        ws.cell(row=1, column=3).value = "keep too"

        expander = ColExpander(ws)
        offset = expander.remove_cols(2, 2)
        assert offset == -1
        assert ws.cell(row=1, column=1).value == "keep"
        assert ws.cell(row=1, column=2).value == "keep too"

    def test_remove_range(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        for i in range(1, 6):
            ws.cell(row=1, column=i).value = f"col{i}"

        expander = ColExpander(ws)
        offset = expander.remove_cols(2, 4)
        assert offset == -3
        assert ws.cell(row=1, column=1).value == "col1"
        assert ws.cell(row=1, column=2).value == "col5"
