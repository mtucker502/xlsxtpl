import openpyxl
import pytest

from xlsxtpl.exceptions import TemplateRenderError, TemplateSyntaxError
from xlsxtpl.jinja_env import create_jinja_env
from xlsxtpl.renderer import SheetRenderer


@pytest.fixture
def env():
    return create_jinja_env()


def make_ws(cells: dict[str, object]) -> openpyxl.worksheet.worksheet.Worksheet:
    """Helper to create a worksheet from a dict of cell coords to values."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for coord, value in cells.items():
        ws[coord] = value
    return ws


class TestSimpleVariables:
    def test_pure_expression_preserves_type(self, env):
        ws = make_ws({"A1": "{{ count }}"})
        SheetRenderer(ws, env).render({"count": 42})
        assert ws["A1"].value == 42

    def test_string_variable(self, env):
        ws = make_ws({"A1": "{{ name }}"})
        SheetRenderer(ws, env).render({"name": "Alice"})
        assert ws["A1"].value == "Alice"

    def test_mixed_content(self, env):
        ws = make_ws({"A1": "Hello {{ name }}!"})
        SheetRenderer(ws, env).render({"name": "Bob"})
        assert ws["A1"].value == "Hello Bob!"

    def test_multiple_expressions_in_cell(self, env):
        ws = make_ws({"A1": "{{ first }} {{ last }}"})
        SheetRenderer(ws, env).render({"first": "Jane", "last": "Doe"})
        assert ws["A1"].value == "Jane Doe"

    def test_non_template_cells_untouched(self, env):
        ws = make_ws({"A1": "plain text", "A2": 42, "A3": "{{ x }}"})
        SheetRenderer(ws, env).render({"x": "ok"})
        assert ws["A1"].value == "plain text"
        assert ws["A2"].value == 42
        assert ws["A3"].value == "ok"

    def test_float_preserved(self, env):
        ws = make_ws({"A1": "{{ price }}"})
        SheetRenderer(ws, env).render({"price": 9.99})
        assert ws["A1"].value == 9.99

    def test_bool_preserved(self, env):
        ws = make_ws({"A1": "{{ flag }}"})
        SheetRenderer(ws, env).render({"flag": True})
        assert ws["A1"].value is True

    def test_none_preserved(self, env):
        ws = make_ws({"A1": "{{ val }}"})
        SheetRenderer(ws, env).render({"val": None})
        assert ws["A1"].value is None


class TestForLoops:
    def test_simple_loop(self, env):
        ws = make_ws({
            "A1": "Header",
            "A2": "{% for item in items %}",
            "A3": "{{ item }}",
            "A4": "{% endfor %}",
            "A5": "Footer",
        })
        SheetRenderer(ws, env).render({"items": ["a", "b", "c"]})

        assert ws["A1"].value == "Header"
        assert ws["A2"].value == "a"
        assert ws["A3"].value == "b"
        assert ws["A4"].value == "c"
        assert ws["A5"].value == "Footer"

    def test_loop_with_multiple_columns(self, env):
        ws = make_ws({
            "A1": "{% for item in items %}",
            "A2": "{{ item.name }}",
            "B2": "{{ item.price }}",
            "A3": "{% endfor %}",
        })
        items = [{"name": "Widget", "price": 10}, {"name": "Gadget", "price": 20}]
        SheetRenderer(ws, env).render({"items": items})

        assert ws["A1"].value == "Widget"
        assert ws["B1"].value == 10
        assert ws["A2"].value == "Gadget"
        assert ws["B2"].value == 20

    def test_empty_loop_removes_block(self, env):
        ws = make_ws({
            "A1": "Header",
            "A2": "{% for item in items %}",
            "A3": "{{ item }}",
            "A4": "{% endfor %}",
            "A5": "Footer",
        })
        SheetRenderer(ws, env).render({"items": []})

        assert ws["A1"].value == "Header"
        assert ws["A2"].value == "Footer"

    def test_single_item_loop(self, env):
        ws = make_ws({
            "A1": "{% for x in items %}",
            "A2": "{{ x }}",
            "A3": "{% endfor %}",
        })
        SheetRenderer(ws, env).render({"items": ["only"]})
        assert ws["A1"].value == "only"

    def test_loop_index(self, env):
        ws = make_ws({
            "A1": "{% for x in items %}",
            "A2": "{{ loop.index }}",
            "B2": "{{ x }}",
            "A3": "{% endfor %}",
        })
        SheetRenderer(ws, env).render({"items": ["a", "b"]})
        assert ws["A1"].value == 1
        assert ws["B1"].value == "a"
        assert ws["A2"].value == 2
        assert ws["B2"].value == "b"

    def test_loop_first_last(self, env):
        ws = make_ws({
            "A1": "{% for x in items %}",
            "A2": "{{ loop.first }}",
            "B2": "{{ loop.last }}",
            "A3": "{% endfor %}",
        })
        SheetRenderer(ws, env).render({"items": ["a", "b", "c"]})
        assert ws["A1"].value is True
        assert ws["B1"].value is False
        assert ws["A2"].value is False
        assert ws["B2"].value is False
        assert ws["A3"].value is False
        assert ws["B3"].value is True


class TestIfBlocks:
    def test_if_true(self, env):
        ws = make_ws({
            "A1": "Header",
            "A2": "{% if show %}",
            "A3": "Visible",
            "A4": "{% endif %}",
            "A5": "Footer",
        })
        SheetRenderer(ws, env).render({"show": True})

        assert ws["A1"].value == "Header"
        assert ws["A2"].value == "Visible"
        assert ws["A3"].value == "Footer"

    def test_if_false(self, env):
        ws = make_ws({
            "A1": "Header",
            "A2": "{% if show %}",
            "A3": "Hidden",
            "A4": "{% endif %}",
            "A5": "Footer",
        })
        SheetRenderer(ws, env).render({"show": False})

        assert ws["A1"].value == "Header"
        assert ws["A2"].value == "Footer"

    def test_if_with_expression_condition(self, env):
        ws = make_ws({
            "A1": "{% if total > 100 %}",
            "A2": "Big order!",
            "A3": "{% endif %}",
        })
        SheetRenderer(ws, env).render({"total": 150})
        assert ws["A1"].value == "Big order!"


class TestNestedBlocks:
    def test_for_inside_for(self, env):
        ws = make_ws({
            "A1": "{% for group in groups %}",
            "A2": "{{ group.name }}",
            "A3": "{% for item in group.items %}",
            "A4": "{{ item }}",
            "A5": "{% endfor %}",
            "A6": "{% endfor %}",
        })
        context = {
            "groups": [
                {"name": "G1", "items": ["a", "b"]},
                {"name": "G2", "items": ["c"]},
            ]
        }
        SheetRenderer(ws, env).render(context)

        # G1 header
        assert ws["A1"].value == "G1"
        # G1 items
        assert ws["A2"].value == "a"
        assert ws["A3"].value == "b"
        # G2 header
        assert ws["A4"].value == "G2"
        # G2 items
        assert ws["A5"].value == "c"

    def test_if_inside_for(self, env):
        ws = make_ws({
            "A1": "{% for item in items %}",
            "A2": "{{ item.name }}",
            "A3": "{% if item.special %}",
            "A4": "SPECIAL",
            "A5": "{% endif %}",
            "A6": "{% endfor %}",
        })
        items = [
            {"name": "A", "special": True},
            {"name": "B", "special": False},
        ]
        SheetRenderer(ws, env).render({"items": items})

        # Item A: name + SPECIAL
        assert ws["A1"].value == "A"
        assert ws["A2"].value == "SPECIAL"
        # Item B: name only (SPECIAL removed)
        assert ws["A3"].value == "B"


class TestColForLoops:
    def test_simple_col_loop(self, env):
        ws = make_ws({
            "A1": "{%col for item in items %}",
            "B1": "{{ item }}",
            "C1": "{%col endfor %}",
        })
        SheetRenderer(ws, env).render({"items": ["a", "b", "c"]})

        # Directive columns removed, 3 body columns remain
        assert ws.cell(row=1, column=1).value == "a"
        assert ws.cell(row=1, column=2).value == "b"
        assert ws.cell(row=1, column=3).value == "c"

    def test_col_loop_multi_column_body(self, env):
        ws = make_ws({
            "A1": "{%col for item in items %}",
            "B1": "{{ item.name }}",
            "C1": "{{ item.value }}",
            "D1": "{%col endfor %}",
        })
        items = [{"name": "X", "value": 10}, {"name": "Y", "value": 20}]
        SheetRenderer(ws, env).render({"items": items})

        # 2 body cols * 2 items = 4 columns
        assert ws.cell(row=1, column=1).value == "X"
        assert ws.cell(row=1, column=2).value == 10
        assert ws.cell(row=1, column=3).value == "Y"
        assert ws.cell(row=1, column=4).value == 20

    def test_col_loop_multi_row(self, env):
        """Column loop with body spanning multiple rows."""
        ws = make_ws({
            "A1": "{%col for q in quarters %}",
            "B1": "{{ q.name }}",
            "B2": "{{ q.revenue }}",
            "C1": "{%col endfor %}",
        })
        quarters = [
            {"name": "Q1", "revenue": 100},
            {"name": "Q2", "revenue": 200},
            {"name": "Q3", "revenue": 300},
        ]
        SheetRenderer(ws, env).render({"quarters": quarters})

        # 3 iterations * 1 body col = 3 columns
        assert ws.cell(row=1, column=1).value == "Q1"
        assert ws.cell(row=2, column=1).value == 100
        assert ws.cell(row=1, column=2).value == "Q2"
        assert ws.cell(row=2, column=2).value == 200
        assert ws.cell(row=1, column=3).value == "Q3"
        assert ws.cell(row=2, column=3).value == 300

    def test_col_loop_empty_list(self, env):
        ws = make_ws({
            "A1": "Before",
            "B1": "{%col for item in items %}",
            "C1": "{{ item }}",
            "D1": "{%col endfor %}",
            "E1": "After",
        })
        SheetRenderer(ws, env).render({"items": []})

        assert ws.cell(row=1, column=1).value == "Before"
        assert ws.cell(row=1, column=2).value == "After"

    def test_col_loop_single_item(self, env):
        ws = make_ws({
            "A1": "{%col for x in items %}",
            "B1": "{{ x }}",
            "C1": "{%col endfor %}",
        })
        SheetRenderer(ws, env).render({"items": ["only"]})
        assert ws.cell(row=1, column=1).value == "only"

    def test_col_loop_index(self, env):
        ws = make_ws({
            "A1": "{%col for x in items %}",
            "B1": "{{ loop.index }}",
            "B2": "{{ x }}",
            "C1": "{%col endfor %}",
        })
        SheetRenderer(ws, env).render({"items": ["a", "b"]})
        assert ws.cell(row=1, column=1).value == 1
        assert ws.cell(row=2, column=1).value == "a"
        assert ws.cell(row=1, column=2).value == 2
        assert ws.cell(row=2, column=2).value == "b"

    def test_col_loop_first_last(self, env):
        ws = make_ws({
            "A1": "{%col for x in items %}",
            "B1": "{{ loop.first }}",
            "B2": "{{ loop.last }}",
            "C1": "{%col endfor %}",
        })
        SheetRenderer(ws, env).render({"items": ["a", "b", "c"]})
        assert ws.cell(row=1, column=1).value is True
        assert ws.cell(row=2, column=1).value is False
        assert ws.cell(row=1, column=2).value is False
        assert ws.cell(row=2, column=2).value is False
        assert ws.cell(row=1, column=3).value is False
        assert ws.cell(row=2, column=3).value is True

    def test_col_loop_preserves_surrounding(self, env):
        """Columns before and after the loop are preserved."""
        ws = make_ws({
            "A1": "Label",
            "B1": "{%col for x in items %}",
            "C1": "{{ x }}",
            "D1": "{%col endfor %}",
            "E1": "Total",
        })
        SheetRenderer(ws, env).render({"items": [1, 2, 3]})

        assert ws.cell(row=1, column=1).value == "Label"
        assert ws.cell(row=1, column=2).value == 1
        assert ws.cell(row=1, column=3).value == 2
        assert ws.cell(row=1, column=4).value == 3
        assert ws.cell(row=1, column=5).value == "Total"


class TestColIfBlocks:
    def test_col_if_true(self, env):
        ws = make_ws({
            "A1": "Before",
            "B1": "{%col if show %}",
            "C1": "Visible",
            "D1": "{%col endif %}",
            "E1": "After",
        })
        SheetRenderer(ws, env).render({"show": True})

        assert ws.cell(row=1, column=1).value == "Before"
        assert ws.cell(row=1, column=2).value == "Visible"
        assert ws.cell(row=1, column=3).value == "After"

    def test_col_if_false(self, env):
        ws = make_ws({
            "A1": "Before",
            "B1": "{%col if show %}",
            "C1": "Hidden",
            "D1": "{%col endif %}",
            "E1": "After",
        })
        SheetRenderer(ws, env).render({"show": False})

        assert ws.cell(row=1, column=1).value == "Before"
        assert ws.cell(row=1, column=2).value == "After"

    def test_col_if_with_expression(self, env):
        ws = make_ws({
            "A1": "{%col if total > 100 %}",
            "B1": "Big!",
            "C1": "{%col endif %}",
        })
        SheetRenderer(ws, env).render({"total": 150})
        assert ws.cell(row=1, column=1).value == "Big!"

    def test_col_if_false_multi_col_body(self, env):
        ws = make_ws({
            "A1": "Keep",
            "B1": "{%col if show %}",
            "C1": "Hide1",
            "D1": "Hide2",
            "E1": "{%col endif %}",
            "F1": "Also keep",
        })
        SheetRenderer(ws, env).render({"show": False})

        assert ws.cell(row=1, column=1).value == "Keep"
        assert ws.cell(row=1, column=2).value == "Also keep"


class TestCrossLoops:
    def test_pivot_table(self, env):
        """Row for inside col for body — the classic pivot table pattern.

        Row directives are in col A (outside the col loop) so they survive
        col-directive removal. The cross-dimensional cell is in the col loop
        body column (C) so it gets duplicated per quarter.
        """
        ws = make_ws({
            "A1": "Metric",
            "B1": "{%col for q in quarters %}",
            "C1": "{{ q.name }}",
            "C3": "{{ data[m.key][q.key] }}",
            "D1": "{%col endfor %}",
            "A2": "{% for m in metrics %}",
            "A3": "{{ m.name }}",
            "A4": "{% endfor %}",
        })
        context = {
            "quarters": [
                {"name": "Q1", "key": "q1"},
                {"name": "Q2", "key": "q2"},
            ],
            "metrics": [
                {"name": "Revenue", "key": "revenue"},
                {"name": "Profit", "key": "profit"},
            ],
            "data": {
                "revenue": {"q1": 100, "q2": 200},
                "profit": {"q1": 10, "q2": 20},
            },
        }
        SheetRenderer(ws, env).render(context)

        # Col A = labels, col B = Q1, col C = Q2
        assert ws.cell(row=1, column=1).value == "Metric"
        assert ws.cell(row=1, column=2).value == "Q1"
        assert ws.cell(row=1, column=3).value == "Q2"
        assert ws.cell(row=2, column=1).value == "Revenue"
        assert ws.cell(row=2, column=2).value == 100
        assert ws.cell(row=2, column=3).value == 200
        assert ws.cell(row=3, column=1).value == "Profit"
        assert ws.cell(row=3, column=2).value == 10
        assert ws.cell(row=3, column=3).value == 20

    def test_loop_and_col_loop_accessible(self, env):
        """Both loop (row) and col_loop variables are accessible in cross cells."""
        ws = make_ws({
            "A1": "Header",
            "B1": "{%col for q in quarters %}",
            "C1": "{{ q }}",
            "C3": "r{{ loop.index }}-c{{ col_loop.index }}",
            "D1": "{%col endfor %}",
            "A2": "{% for m in metrics %}",
            "A3": "{{ m }}",
            "A4": "{% endfor %}",
        })
        context = {
            "quarters": ["Q1", "Q2"],
            "metrics": ["Rev", "Cost"],
        }
        SheetRenderer(ws, env).render(context)

        # Row 1: headers
        assert ws.cell(row=1, column=1).value == "Header"
        assert ws.cell(row=1, column=2).value == "Q1"
        assert ws.cell(row=1, column=3).value == "Q2"
        # Row 2: first metric (loop.index=1, col_loop.index=1 or 2)
        assert ws.cell(row=2, column=1).value == "Rev"
        assert ws.cell(row=2, column=2).value == "r1-c1"
        assert ws.cell(row=2, column=3).value == "r1-c2"
        # Row 3: second metric (loop.index=2)
        assert ws.cell(row=3, column=1).value == "Cost"
        assert ws.cell(row=3, column=2).value == "r2-c1"
        assert ws.cell(row=3, column=3).value == "r2-c2"

    def test_col_for_and_row_for_independent(self, env):
        """Col for + row for on same sheet, no cross-reference needed."""
        ws = make_ws({
            "A1": "Header",
            "B1": "{%col for q in quarters %}",
            "C1": "{{ q }}",
            "D1": "{%col endfor %}",
            "A2": "{% for m in metrics %}",
            "A3": "{{ m }}",
            "A4": "{% endfor %}",
        })
        context = {
            "quarters": ["Q1", "Q2", "Q3"],
            "metrics": ["Rev", "Cost"],
        }
        SheetRenderer(ws, env).render(context)

        # Col-only row
        assert ws.cell(row=1, column=1).value == "Header"
        assert ws.cell(row=1, column=2).value == "Q1"
        assert ws.cell(row=1, column=3).value == "Q2"
        assert ws.cell(row=1, column=4).value == "Q3"
        # Row-only rows
        assert ws.cell(row=2, column=1).value == "Rev"
        assert ws.cell(row=3, column=1).value == "Cost"

    def test_empty_col_loop_with_row_loop(self, env):
        """Empty col loop doesn't break row loop processing."""
        ws = make_ws({
            "A1": "Header",
            "B1": "{%col for q in quarters %}",
            "C1": "{{ q }}",
            "D1": "{%col endfor %}",
            "A2": "{% for m in metrics %}",
            "A3": "{{ m }}",
            "A4": "{% endfor %}",
        })
        context = {
            "quarters": [],
            "metrics": ["Rev", "Cost"],
        }
        SheetRenderer(ws, env).render(context)

        # Col loop removed entirely, row loop still works
        assert ws.cell(row=1, column=1).value == "Header"
        assert ws.cell(row=2, column=1).value == "Rev"
        assert ws.cell(row=3, column=1).value == "Cost"

    def test_col_if_with_row_for(self, env):
        """Col if alongside row for — regression test."""
        ws = make_ws({
            "A1": "Label",
            "B1": "{%col if show_extra %}",
            "C1": "Extra",
            "D1": "{%col endif %}",
            "A2": "{% for m in metrics %}",
            "A3": "{{ m }}",
            "A4": "{% endfor %}",
        })
        context = {
            "show_extra": True,
            "metrics": ["Rev", "Cost"],
        }
        SheetRenderer(ws, env).render(context)

        assert ws.cell(row=1, column=1).value == "Label"
        assert ws.cell(row=1, column=2).value == "Extra"
        assert ws.cell(row=2, column=1).value == "Rev"
        assert ws.cell(row=3, column=1).value == "Cost"


class TestColSyntaxErrors:
    def test_mismatched_col_tags(self, env):
        ws = make_ws({
            "A1": "{%col for x in items %}",
            "B1": "{{ x }}",
            "C1": "{%col endif %}",
        })
        with pytest.raises(TemplateSyntaxError, match="Mismatched column"):
            SheetRenderer(ws, env).render({"items": []})

    def test_unclosed_col_block(self, env):
        ws = make_ws({
            "A1": "{%col for x in items %}",
            "B1": "{{ x }}",
        })
        with pytest.raises(TemplateSyntaxError, match="Unclosed"):
            SheetRenderer(ws, env).render({"items": []})

    def test_bad_col_iterable(self, env):
        ws = make_ws({
            "A1": "{%col for x in nonexistent %}",
            "B1": "{{ x }}",
            "C1": "{%col endfor %}",
        })
        with pytest.raises(TemplateRenderError, match="Failed to evaluate"):
            SheetRenderer(ws, env).render({})


class TestSyntaxErrors:
    def test_mismatched_tags(self, env):
        ws = make_ws({
            "A1": "{% for x in items %}",
            "A2": "{{ x }}",
            "A3": "{% endif %}",  # wrong closer
        })
        with pytest.raises(TemplateSyntaxError, match="Mismatched"):
            SheetRenderer(ws, env).render({"items": []})

    def test_unclosed_block(self, env):
        ws = make_ws({
            "A1": "{% for x in items %}",
            "A2": "{{ x }}",
        })
        with pytest.raises(TemplateSyntaxError, match="Unclosed"):
            SheetRenderer(ws, env).render({"items": []})

    def test_bad_iterable(self, env):
        ws = make_ws({
            "A1": "{% for x in nonexistent %}",
            "A2": "{{ x }}",
            "A3": "{% endfor %}",
        })
        with pytest.raises(TemplateRenderError, match="Failed to evaluate"):
            SheetRenderer(ws, env).render({})

    def test_bad_expression(self, env):
        ws = make_ws({"A1": "{{ undefined_var }}"})
        with pytest.raises(TemplateRenderError):
            SheetRenderer(ws, env).render({})
