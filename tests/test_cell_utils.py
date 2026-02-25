from datetime import datetime

import openpyxl
import pytest

from openpyxl.utils import get_column_letter

from xlsxtpl.cell_utils import (
    cell_has_only_expression,
    coerce_type,
    copy_cell,
    copy_column_dimensions,
    copy_row_dimensions,
    extract_block_directive,
    extract_col_block_directive,
    has_template_tag,
    is_block_tag,
    is_col_block_tag,
)


class TestCellHasOnlyExpression:
    def test_simple_expression(self):
        assert cell_has_only_expression("{{ name }}") is not None
        assert cell_has_only_expression("{{ name }}").group(1) == "name"

    def test_expression_with_filter(self):
        m = cell_has_only_expression("{{ price | round(2) }}")
        assert m is not None
        assert m.group(1) == "price | round(2)"

    def test_expression_with_whitespace(self):
        assert cell_has_only_expression("  {{ x }}  ") is not None

    def test_mixed_content_not_pure(self):
        assert cell_has_only_expression("Total: {{ x }}") is None

    def test_non_string(self):
        assert cell_has_only_expression(42) is None
        assert cell_has_only_expression(None) is None

    def test_block_tag_not_expression(self):
        assert cell_has_only_expression("{% for x in items %}") is None


class TestHasTemplateTag:
    def test_expression_tag(self):
        assert has_template_tag("{{ name }}")

    def test_block_tag(self):
        assert has_template_tag("{% for x in items %}")

    def test_mixed(self):
        assert has_template_tag("Hello {{ name }}!")

    def test_plain_text(self):
        assert not has_template_tag("Hello World")

    def test_non_string(self):
        assert not has_template_tag(42)
        assert not has_template_tag(None)


class TestIsBlockTag:
    def test_for_tag(self):
        assert is_block_tag("{% for x in items %}")

    def test_endfor_tag(self):
        assert is_block_tag("{% endfor %}")

    def test_if_tag(self):
        assert is_block_tag("{% if show %}")

    def test_endif_tag(self):
        assert is_block_tag("{% endif %}")

    def test_expression_not_block(self):
        assert not is_block_tag("{{ name }}")

    def test_plain_text(self):
        assert not is_block_tag("hello")

    def test_non_string(self):
        assert not is_block_tag(42)


class TestExtractBlockDirective:
    def test_for_directive(self):
        result = extract_block_directive("{% for item in items %}")
        assert result == {"type": "for", "var": "item", "iterable": "items"}

    def test_for_with_complex_iterable(self):
        result = extract_block_directive("{% for p in data.products %}")
        assert result["type"] == "for"
        assert result["var"] == "p"
        assert result["iterable"] == "data.products"

    def test_endfor(self):
        result = extract_block_directive("{% endfor %}")
        assert result == {"type": "endfor"}

    def test_if_directive(self):
        result = extract_block_directive("{% if show_total %}")
        assert result == {"type": "if", "condition": "show_total"}

    def test_if_complex_condition(self):
        result = extract_block_directive("{% if total > 100 %}")
        assert result["type"] == "if"
        assert result["condition"] == "total > 100"

    def test_endif(self):
        result = extract_block_directive("{% endif %}")
        assert result == {"type": "endif"}

    def test_not_a_block(self):
        assert extract_block_directive("{{ name }}") is None
        assert extract_block_directive("hello") is None

    def test_with_dash_whitespace_control(self):
        result = extract_block_directive("{%- for x in items -%}")
        assert result is not None
        assert result["type"] == "for"


class TestCoerceType:
    def test_int_from_int_original(self):
        assert coerce_type("42", 0) == 42
        assert isinstance(coerce_type("42", 0), int)

    def test_float_from_float_original(self):
        assert coerce_type("3.14", 0.0) == 3.14
        assert isinstance(coerce_type("3.14", 0.0), float)

    def test_float_string_from_int_original(self):
        result = coerce_type("3.14", 0)
        assert result == 3.14

    def test_non_numeric_passthrough(self):
        assert coerce_type("hello", "original") == "hello"

    def test_empty_string(self):
        assert coerce_type("", 0) == ""

    def test_bool_true(self):
        assert coerce_type("True", True) is True

    def test_bool_false(self):
        assert coerce_type("False", False) is False

    def test_date_coercion(self):
        result = coerce_type("2024-01-15", datetime.now())
        assert isinstance(result, datetime)


class TestCopyCell:
    def test_copies_value(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "hello"
        ws["B1"] = None
        copy_cell(ws["A1"], ws["B1"])
        assert ws["B1"].value == "hello"

    def test_copies_style(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"].font = openpyxl.styles.Font(bold=True)
        ws["A1"].value = "bold"
        copy_cell(ws["A1"], ws["B1"])
        assert ws["B1"].font.bold is True


class TestCopyRowDimensions:
    def test_copies_height(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.row_dimensions[1].height = 30
        copy_row_dimensions(ws, 1, 2)
        assert ws.row_dimensions[2].height == 30

    def test_no_error_on_missing_source(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        # Should not raise even if row 99 has no explicit dimensions
        copy_row_dimensions(ws, 99, 1)


class TestIsColBlockTag:
    def test_col_for_tag(self):
        assert is_col_block_tag("{%col for x in items %}")

    def test_col_endfor_tag(self):
        assert is_col_block_tag("{%col endfor %}")

    def test_col_if_tag(self):
        assert is_col_block_tag("{%col if show %}")

    def test_col_endif_tag(self):
        assert is_col_block_tag("{%col endif %}")

    def test_regular_block_not_col(self):
        assert not is_col_block_tag("{% for x in items %}")

    def test_expression_not_col_block(self):
        assert not is_col_block_tag("{{ name }}")

    def test_plain_text(self):
        assert not is_col_block_tag("hello")

    def test_non_string(self):
        assert not is_col_block_tag(42)


class TestExtractColBlockDirective:
    def test_col_for_directive(self):
        result = extract_col_block_directive("{%col for item in items %}")
        assert result == {"type": "for", "var": "item", "iterable": "items"}

    def test_col_for_complex_iterable(self):
        result = extract_col_block_directive("{%col for p in data.products %}")
        assert result["type"] == "for"
        assert result["var"] == "p"
        assert result["iterable"] == "data.products"

    def test_col_endfor(self):
        result = extract_col_block_directive("{%col endfor %}")
        assert result == {"type": "endfor"}

    def test_col_if_directive(self):
        result = extract_col_block_directive("{%col if show_total %}")
        assert result == {"type": "if", "condition": "show_total"}

    def test_col_if_complex_condition(self):
        result = extract_col_block_directive("{%col if total > 100 %}")
        assert result["type"] == "if"
        assert result["condition"] == "total > 100"

    def test_col_endif(self):
        result = extract_col_block_directive("{%col endif %}")
        assert result == {"type": "endif"}

    def test_not_a_col_block(self):
        assert extract_col_block_directive("{{ name }}") is None
        assert extract_col_block_directive("{% for x in items %}") is None
        assert extract_col_block_directive("hello") is None


class TestCopyColumnDimensions:
    def test_copies_width(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.column_dimensions["A"].width = 30
        copy_column_dimensions(ws, 1, 2)
        assert ws.column_dimensions[get_column_letter(2)].width == 30

    def test_no_error_on_missing_source(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        # Should not raise even if col 99 has no explicit dimensions
        copy_column_dimensions(ws, 99, 1)
