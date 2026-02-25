from __future__ import annotations

import tempfile
from pathlib import Path

import openpyxl
import pytest


@pytest.fixture
def tmp_dir():
    """Provide a temporary directory for test outputs."""
    with tempfile.TemporaryDirectory() as d:
        yield Path(d)


@pytest.fixture
def make_workbook():
    """Factory fixture to create an in-memory openpyxl workbook."""
    def _make():
        return openpyxl.Workbook()
    return _make


@pytest.fixture
def make_template_file(tmp_dir):
    """Factory fixture: creates a template xlsx file from cell data.

    Usage:
        path = make_template_file({
            "A1": "Hello {{ name }}",
            "B1": 42,
        })
    """
    def _make(cells: dict[str, object], sheet_name: str | None = None) -> Path:
        wb = openpyxl.Workbook()
        ws = wb.active
        if sheet_name:
            ws.title = sheet_name
        for coord, value in cells.items():
            ws[coord] = value
        path = tmp_dir / "template.xlsx"
        wb.save(str(path))
        return path
    return _make
