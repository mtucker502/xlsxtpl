import openpyxl
import pytest

from xlsxtpl.row_expander import RowExpander


@pytest.fixture
def ws_with_body():
    """Workbook with a 3-row body (rows 2-4) surrounded by header/footer."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Header"
    ws["A2"] = "Row A"
    ws["B2"] = 10
    ws["A3"] = "Row B"
    ws["B3"] = 20
    ws["A4"] = "Row C"
    ws["B4"] = 30
    ws["A5"] = "Footer"
    # Set a custom height on row 2
    ws.row_dimensions[2].height = 25
    return ws


class TestExpandForLoop:
    def test_single_iteration_no_change(self, ws_with_body):
        expander = RowExpander(ws_with_body)
        added = expander.expand_for_loop(body_start=2, body_end=4, iteration_count=1)
        assert added == 0
        assert ws_with_body.max_row == 5

    def test_zero_iterations_no_change(self, ws_with_body):
        expander = RowExpander(ws_with_body)
        added = expander.expand_for_loop(body_start=2, body_end=4, iteration_count=0)
        assert added == 0

    def test_two_iterations(self, ws_with_body):
        ws = ws_with_body
        expander = RowExpander(ws)
        added = expander.expand_for_loop(body_start=2, body_end=4, iteration_count=2)
        assert added == 3  # (2-1) * 3 = 3

        # Original body preserved in rows 2-4
        assert ws["A2"].value == "Row A"
        assert ws["A3"].value == "Row B"
        assert ws["A4"].value == "Row C"

        # Copy in rows 5-7
        assert ws["A5"].value == "Row A"
        assert ws["B5"].value == 10
        assert ws["A6"].value == "Row B"
        assert ws["A7"].value == "Row C"

        # Footer shifted to row 8
        assert ws["A8"].value == "Footer"

    def test_three_iterations(self, ws_with_body):
        ws = ws_with_body
        expander = RowExpander(ws)
        added = expander.expand_for_loop(body_start=2, body_end=4, iteration_count=3)
        assert added == 6  # (3-1) * 3 = 6

        # Third iteration in rows 8-10
        assert ws["A8"].value == "Row A"
        assert ws["A9"].value == "Row B"
        assert ws["A10"].value == "Row C"

        # Footer at row 11
        assert ws["A11"].value == "Footer"

    def test_style_preservation(self, ws_with_body):
        ws = ws_with_body
        ws["A2"].font = openpyxl.styles.Font(bold=True)
        expander = RowExpander(ws)
        expander.expand_for_loop(body_start=2, body_end=4, iteration_count=2)
        # Copied row should have bold font
        assert ws["A5"].font.bold is True

    def test_row_height_preservation(self, ws_with_body):
        ws = ws_with_body
        expander = RowExpander(ws)
        expander.expand_for_loop(body_start=2, body_end=4, iteration_count=2)
        # Row 5 (first row of iteration 2) should have same height as row 2
        assert ws.row_dimensions[5].height == 25


class TestRemoveRows:
    def test_remove_single_row(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "keep"
        ws["A2"] = "delete"
        ws["A3"] = "keep too"

        expander = RowExpander(ws)
        offset = expander.remove_rows(2, 2)
        assert offset == -1
        assert ws["A1"].value == "keep"
        assert ws["A2"].value == "keep too"

    def test_remove_range(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        for i in range(1, 6):
            ws[f"A{i}"] = f"row{i}"

        expander = RowExpander(ws)
        offset = expander.remove_rows(2, 4)
        assert offset == -3
        assert ws["A1"].value == "row1"
        assert ws["A2"].value == "row5"
