import openpyxl
import pytest

from xlsxtpl import XlsxTemplate


class TestEndToEnd:
    def test_simple_render(self, make_template_file, tmp_dir):
        path = make_template_file({
            "A1": "Name:",
            "B1": "{{ name }}",
            "A2": "Age:",
            "B2": "{{ age }}",
        })
        tpl = XlsxTemplate(path)
        tpl.render({"name": "Alice", "age": 30})

        output = tmp_dir / "output.xlsx"
        tpl.save(output)

        wb = openpyxl.load_workbook(str(output))
        ws = wb.active
        assert ws["B1"].value == "Alice"
        assert ws["B2"].value == 30

    def test_for_loop_render(self, make_template_file, tmp_dir):
        path = make_template_file({
            "A1": "Items",
            "A2": "{% for item in items %}",
            "A3": "{{ item.name }}",
            "B3": "{{ item.qty }}",
            "A4": "{% endfor %}",
            "A5": "Total: {{ total }}",
        })
        tpl = XlsxTemplate(path)
        items = [
            {"name": "Widget", "qty": 5},
            {"name": "Gadget", "qty": 3},
            {"name": "Doohickey", "qty": 1},
        ]
        tpl.render({"items": items, "total": 9})

        output = tmp_dir / "output.xlsx"
        tpl.save(output)

        wb = openpyxl.load_workbook(str(output))
        ws = wb.active
        assert ws["A1"].value == "Items"
        assert ws["A2"].value == "Widget"
        assert ws["B2"].value == 5
        assert ws["A3"].value == "Gadget"
        assert ws["B3"].value == 3
        assert ws["A4"].value == "Doohickey"
        assert ws["B4"].value == 1
        assert ws["A5"].value == "Total: 9"

    def test_if_block_render(self, make_template_file, tmp_dir):
        path = make_template_file({
            "A1": "Report",
            "A2": "{% if show_detail %}",
            "A3": "Detail: {{ detail }}",
            "A4": "{% endif %}",
            "A5": "End",
        })
        tpl = XlsxTemplate(path)
        tpl.render({"show_detail": True, "detail": "important info"})

        output = tmp_dir / "output.xlsx"
        tpl.save(output)

        wb = openpyxl.load_workbook(str(output))
        ws = wb.active
        assert ws["A1"].value == "Report"
        assert ws["A2"].value == "Detail: important info"
        assert ws["A3"].value == "End"

    def test_style_preserved_through_render(self, tmp_dir):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "{{ name }}"
        ws["A1"].font = openpyxl.styles.Font(bold=True, size=14)
        tpl_path = tmp_dir / "styled.xlsx"
        wb.save(str(tpl_path))

        tpl = XlsxTemplate(tpl_path)
        tpl.render({"name": "Styled"})

        output = tmp_dir / "output.xlsx"
        tpl.save(output)

        wb2 = openpyxl.load_workbook(str(output))
        cell = wb2.active["A1"]
        assert cell.value == "Styled"
        assert cell.font.bold is True
        assert cell.font.size == 14

    def test_custom_filter(self, make_template_file, tmp_dir):
        path = make_template_file({"A1": "{{ price | currency }}"})
        tpl = XlsxTemplate(path)
        tpl.jinja_env.filters["currency"] = lambda v: f"${v:,.2f}"
        tpl.render({"price": 1234.5})

        output = tmp_dir / "output.xlsx"
        tpl.save(output)

        wb = openpyxl.load_workbook(str(output))
        assert wb.active["A1"].value == "$1,234.50"

    def test_multiple_sheets(self, tmp_dir):
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1["A1"] = "{{ x }}"
        ws2 = wb.create_sheet("Sheet2")
        ws2["A1"] = "{{ y }}"
        tpl_path = tmp_dir / "multi.xlsx"
        wb.save(str(tpl_path))

        tpl = XlsxTemplate(tpl_path)
        tpl.render({"x": "hello", "y": "world"})

        output = tmp_dir / "output.xlsx"
        tpl.save(output)

        wb2 = openpyxl.load_workbook(str(output))
        assert wb2["Sheet1"]["A1"].value == "hello"
        assert wb2["Sheet2"]["A1"].value == "world"

    def test_type_preservation_int(self, make_template_file, tmp_dir):
        path = make_template_file({"A1": "{{ count }}"})
        tpl = XlsxTemplate(path)
        tpl.render({"count": 42})

        output = tmp_dir / "output.xlsx"
        tpl.save(output)

        wb = openpyxl.load_workbook(str(output))
        val = wb.active["A1"].value
        assert val == 42
        assert isinstance(val, int)

    def test_empty_template(self, make_template_file, tmp_dir):
        path = make_template_file({"A1": "No templates here"})
        tpl = XlsxTemplate(path)
        tpl.render({"anything": "ignored"})

        output = tmp_dir / "output.xlsx"
        tpl.save(output)

        wb = openpyxl.load_workbook(str(output))
        assert wb.active["A1"].value == "No templates here"
