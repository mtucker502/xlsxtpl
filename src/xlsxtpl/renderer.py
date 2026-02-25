from __future__ import annotations

from dataclasses import dataclass, field
from typing import Any

import jinja2
from jinja2 import Undefined
from openpyxl.worksheet.worksheet import Worksheet

from .cell_utils import (
    cell_has_only_expression,
    extract_block_directive,
    extract_col_block_directive,
    has_template_tag,
    is_block_tag,
    is_col_block_tag,
)
from .col_expander import ColExpander
from .exceptions import TemplateRenderError, TemplateSyntaxError
from .row_expander import RowExpander


@dataclass
class RowBlock:
    """Represents a matched block (for-loop or if-block) with row boundaries."""

    block_type: str  # "for" or "if"
    open_row: int  # row of {% for/if %}
    close_row: int  # row of {% endfor/endif %}
    directive: dict[str, Any] = field(default_factory=dict)
    children: list[RowBlock] = field(default_factory=list)

    @property
    def body_start(self) -> int:
        return self.open_row + 1

    @property
    def body_end(self) -> int:
        return self.close_row - 1

    @property
    def body_row_count(self) -> int:
        return self.close_row - self.open_row - 1


@dataclass
class ColBlock:
    """Represents a matched block (for-loop or if-block) with column boundaries."""

    block_type: str  # "for" or "if"
    open_col: int  # column of {%col for/if %}
    close_col: int  # column of {%col endfor/endif %}
    directive: dict[str, Any] = field(default_factory=dict)
    children: list[ColBlock] = field(default_factory=list)

    @property
    def body_start(self) -> int:
        return self.open_col + 1

    @property
    def body_end(self) -> int:
        return self.close_col - 1

    @property
    def body_col_count(self) -> int:
        return self.close_col - self.open_col - 1


class SheetRenderer:
    """Renders a single worksheet by processing block directives and expressions."""

    def __init__(self, ws: Worksheet, env: jinja2.Environment) -> None:
        self.ws = ws
        self.env = env
        self.expander = RowExpander(ws)
        self.col_expander = ColExpander(ws)
        self._col_context: dict[int, dict[str, Any]] = {}

    def render(self, context: dict[str, Any]) -> None:
        """Full rendering pipeline for the worksheet."""
        # Process column blocks first (before row blocks)
        col_blocks = self._scan_col_blocks_in_range(1, self.ws.max_column or 0)
        self._process_col_blocks(col_blocks, context)

        # Then process row blocks
        blocks = self._scan_blocks_in_range(1, self.ws.max_row or 0)
        self._process_blocks(blocks, context)
        self._render_remaining_cells(context)

    # --- Column context helpers ---

    def _merge_col_context(self, col: int, context: dict[str, Any]) -> dict[str, Any]:
        """Merge stored column context for a column into the given context.

        Caller's context wins on conflict, so row-loop variables take
        precedence over column-loop variables.
        """
        col_ctx = self._col_context.get(col)
        if col_ctx is None:
            return context
        return {**col_ctx, **context}

    def _shift_col_context(self, at_col: int, delta: int) -> None:
        """Shift column context entries after a structural column change.

        Positive delta: entries >= at_col shift right.
        Negative delta: entries in the removed range are dropped,
        entries above shift down.
        """
        if delta == 0:
            return
        old = self._col_context
        new: dict[int, dict[str, Any]] = {}
        if delta > 0:
            for col, ctx in old.items():
                if col >= at_col:
                    new[col + delta] = ctx
                else:
                    new[col] = ctx
        else:
            # delta < 0 — columns [at_col, at_col - delta) were removed
            removed_end = at_col - delta  # exclusive
            for col, ctx in old.items():
                if at_col <= col < removed_end:
                    continue  # dropped
                elif col >= removed_end:
                    new[col + delta] = ctx
                else:
                    new[col] = ctx
        self._col_context = new

    def _store_col_context(
        self, start_col: int, end_col: int, context: dict[str, Any]
    ) -> None:
        """Store per-column context for deferred rendering.

        If a column already has context (from an inner loop), the existing
        context takes precedence (inner loop wins).
        """
        for col in range(start_col, end_col + 1):
            existing = self._col_context.get(col)
            if existing is not None:
                self._col_context[col] = {**context, **existing}
            else:
                self._col_context[col] = context

    # --- Phase A+B: Scan and match blocks ---

    def _scan_blocks_in_range(self, start_row: int, end_row: int) -> list[RowBlock]:
        """Scan a row range for block directives and match them into a tree."""
        openers: list[tuple[int, dict[str, Any]]] = []
        closers: list[tuple[int, str]] = []

        for row in range(start_row, end_row + 1):
            for col in range(1, (self.ws.max_column or 1) + 1):
                val = self.ws.cell(row=row, column=col).value
                if not is_block_tag(val):
                    continue
                directive = extract_block_directive(val)
                if directive is None:
                    continue
                dtype = directive["type"]
                if dtype in ("for", "if"):
                    openers.append((row, directive))
                elif dtype in ("endfor", "endif"):
                    closers.append((row, dtype))
                break  # only one directive per row

        return self._match_blocks(openers, closers)

    def _match_blocks(
        self,
        openers: list[tuple[int, dict[str, Any]]],
        closers: list[tuple[int, str]],
    ) -> list[RowBlock]:
        """Stack-based matching of openers to closers, building a tree."""
        events: list[tuple[int, str, dict[str, Any] | None]] = []
        for row, directive in openers:
            events.append((row, "open", directive))
        for row, dtype in closers:
            events.append((row, "close", {"type": dtype}))
        events.sort(key=lambda e: e[0])

        stack: list[RowBlock] = []
        top_level: list[RowBlock] = []

        for row, kind, directive in events:
            if kind == "open":
                block = RowBlock(
                    block_type=directive["type"],
                    open_row=row,
                    close_row=-1,
                    directive=directive,
                )
                stack.append(block)
            elif kind == "close":
                expected_close = "end" + (stack[-1].block_type if stack else "???")
                actual_close = directive["type"]
                if not stack or actual_close != expected_close:
                    raise TemplateSyntaxError(
                        f"Mismatched block tag at row {row}: "
                        f"found {{% {actual_close} %}} but expected {{% {expected_close} %}}"
                    )
                block = stack.pop()
                block.close_row = row
                if stack:
                    stack[-1].children.append(block)
                else:
                    top_level.append(block)

        if stack:
            unmatched = stack[-1]
            raise TemplateSyntaxError(
                f"Unclosed {{% {unmatched.block_type} %}} at row {unmatched.open_row}"
            )

        return top_level

    # --- Phase C: Process blocks ---

    def _process_blocks(
        self, blocks: list[RowBlock], context: dict[str, Any]
    ) -> int:
        """Process blocks bottom-up at this level. Returns total row offset."""
        total_delta = 0
        # Process bottom-up so lower block changes don't affect upper block row numbers
        for block in reversed(blocks):
            if block.block_type == "for":
                delta = self._process_for_block(block, context)
            elif block.block_type == "if":
                delta = self._process_if_block(block, context)
            else:
                delta = 0
            total_delta += delta
        return total_delta

    def _process_for_block(self, block: RowBlock, context: dict[str, Any]) -> int:
        """Expand a for-loop block. Returns net row change."""
        var_name = block.directive["var"]
        iterable_expr = block.directive["iterable"]

        try:
            expr = self.env.compile_expression(iterable_expr)
            items = list(expr(**context))
        except Exception as e:
            raise TemplateRenderError(
                f"Failed to evaluate iterable '{iterable_expr}' at row {block.open_row}: {e}"
            ) from e

        iteration_count = len(items)
        body_row_count = block.body_row_count

        if body_row_count <= 0:
            delta = self.expander.remove_rows(block.open_row, block.close_row)
            return delta

        if iteration_count == 0:
            delta = self.expander.remove_rows(block.open_row, block.close_row)
            return delta

        # Expand body rows for all iterations
        rows_added = self.expander.expand_for_loop(
            block.body_start, block.body_end, iteration_count
        )

        # Process each iteration, tracking cumulative row changes from nested blocks
        cumulative_child_delta = 0
        for i, item in enumerate(items):
            iter_start = block.body_start + i * body_row_count + cumulative_child_delta
            iter_end = iter_start + body_row_count - 1

            loop_context = {
                **context,
                var_name: item,
                "loop": {
                    "index": i + 1,
                    "index0": i,
                    "first": i == 0,
                    "last": i == iteration_count - 1,
                    "length": iteration_count,
                    "revindex": iteration_count - i,
                    "revindex0": iteration_count - i - 1,
                },
            }

            # Re-scan this iteration's range for nested blocks
            nested_blocks = self._scan_blocks_in_range(iter_start, iter_end)
            child_delta = self._process_blocks(nested_blocks, loop_context)
            cumulative_child_delta += child_delta

            # Render expression cells (adjusted for child block changes)
            self._render_row_range(iter_start, iter_end + child_delta, loop_context)

        # Delete the {% endfor %} row (shifted by expansion + child changes)
        new_close_row = block.close_row + rows_added + cumulative_child_delta
        self.expander.remove_rows(new_close_row, new_close_row)

        # Delete the {% for %} row
        self.expander.remove_rows(block.open_row, block.open_row)

        return rows_added + cumulative_child_delta - 2

    def _process_if_block(self, block: RowBlock, context: dict[str, Any]) -> int:
        """Process an if-block. Returns net row change."""
        condition_expr = block.directive["condition"]

        try:
            expr = self.env.compile_expression(condition_expr)
            result = expr(**context)
        except Exception as e:
            raise TemplateRenderError(
                f"Failed to evaluate condition '{condition_expr}' at row {block.open_row}: {e}"
            ) from e

        if result:
            # Condition is true: keep body, remove directive rows
            # Re-scan body for nested blocks
            nested_blocks = self._scan_blocks_in_range(block.body_start, block.body_end)
            child_delta = self._process_blocks(nested_blocks, context)

            # Render body cells (adjusted for child changes)
            self._render_row_range(block.body_start, block.body_end + child_delta, context)

            # Remove close row first (higher number), then open row
            adjusted_close = block.close_row + child_delta
            self.expander.remove_rows(adjusted_close, adjusted_close)
            self.expander.remove_rows(block.open_row, block.open_row)
            return child_delta - 2
        else:
            # Condition is false: remove entire block
            delta = self.expander.remove_rows(block.open_row, block.close_row)
            return delta

    # --- Column blocks ---

    def _scan_col_blocks_in_range(
        self, start_col: int, end_col: int
    ) -> list[ColBlock]:
        """Scan a column range for column block directives and match them."""
        openers: list[tuple[int, dict[str, Any]]] = []
        closers: list[tuple[int, str]] = []

        for col in range(start_col, end_col + 1):
            for row in range(1, (self.ws.max_row or 1) + 1):
                val = self.ws.cell(row=row, column=col).value
                if not is_col_block_tag(val):
                    continue
                directive = extract_col_block_directive(val)
                if directive is None:
                    continue
                dtype = directive["type"]
                if dtype in ("for", "if"):
                    openers.append((col, directive))
                elif dtype in ("endfor", "endif"):
                    closers.append((col, dtype))
                break  # only one directive per column

        return self._match_col_blocks(openers, closers)

    def _match_col_blocks(
        self,
        openers: list[tuple[int, dict[str, Any]]],
        closers: list[tuple[int, str]],
    ) -> list[ColBlock]:
        """Stack-based matching of column openers to closers, building a tree."""
        events: list[tuple[int, str, dict[str, Any] | None]] = []
        for col, directive in openers:
            events.append((col, "open", directive))
        for col, dtype in closers:
            events.append((col, "close", {"type": dtype}))
        events.sort(key=lambda e: e[0])

        stack: list[ColBlock] = []
        top_level: list[ColBlock] = []

        for col, kind, directive in events:
            if kind == "open":
                block = ColBlock(
                    block_type=directive["type"],
                    open_col=col,
                    close_col=-1,
                    directive=directive,
                )
                stack.append(block)
            elif kind == "close":
                expected_close = "end" + (stack[-1].block_type if stack else "???")
                actual_close = directive["type"]
                if not stack or actual_close != expected_close:
                    raise TemplateSyntaxError(
                        f"Mismatched column block tag at column {col}: "
                        f"found {{%col {actual_close} %}} but expected "
                        f"{{%col {expected_close} %}}"
                    )
                block = stack.pop()
                block.close_col = col
                if stack:
                    stack[-1].children.append(block)
                else:
                    top_level.append(block)

        if stack:
            unmatched = stack[-1]
            raise TemplateSyntaxError(
                f"Unclosed {{%col {unmatched.block_type} %}} at column "
                f"{unmatched.open_col}"
            )

        return top_level

    def _process_col_blocks(
        self, blocks: list[ColBlock], context: dict[str, Any]
    ) -> int:
        """Process column blocks right-to-left. Returns total column offset."""
        total_delta = 0
        for block in reversed(blocks):
            if block.block_type == "for":
                delta = self._process_col_for_block(block, context)
            elif block.block_type == "if":
                delta = self._process_col_if_block(block, context)
            else:
                delta = 0
            total_delta += delta
        return total_delta

    def _process_col_for_block(
        self, block: ColBlock, context: dict[str, Any]
    ) -> int:
        """Expand a column for-loop block. Returns net column change."""
        var_name = block.directive["var"]
        iterable_expr = block.directive["iterable"]

        try:
            expr = self.env.compile_expression(iterable_expr)
            items = list(expr(**context))
        except Exception as e:
            raise TemplateRenderError(
                f"Failed to evaluate iterable '{iterable_expr}' "
                f"at column {block.open_col}: {e}"
            ) from e

        iteration_count = len(items)
        body_col_count = block.body_col_count

        if body_col_count <= 0:
            delta = self.col_expander.remove_cols(block.open_col, block.close_col)
            self._shift_col_context(block.open_col, delta)
            return delta

        if iteration_count == 0:
            delta = self.col_expander.remove_cols(block.open_col, block.close_col)
            self._shift_col_context(block.open_col, delta)
            return delta

        # Expand body columns for all iterations
        cols_added = self.col_expander.expand_for_loop(
            block.body_start, block.body_end, iteration_count
        )
        self._shift_col_context(block.body_end + 1, cols_added)

        # Process each iteration
        cumulative_child_delta = 0
        for i, item in enumerate(items):
            iter_start = block.body_start + i * body_col_count + cumulative_child_delta
            iter_end = iter_start + body_col_count - 1

            loop_info = {
                "index": i + 1,
                "index0": i,
                "first": i == 0,
                "last": i == iteration_count - 1,
                "length": iteration_count,
                "revindex": iteration_count - i,
                "revindex0": iteration_count - i - 1,
            }
            loop_context = {
                **context,
                var_name: item,
                "loop": loop_info,
                "col_loop": loop_info,
            }

            # Re-scan this iteration's range for nested column blocks
            nested_blocks = self._scan_col_blocks_in_range(iter_start, iter_end)
            child_delta = self._process_col_blocks(nested_blocks, loop_context)
            cumulative_child_delta += child_delta

            # Store context per column for deferred rendering
            self._store_col_context(
                iter_start, iter_end + child_delta, loop_context
            )

        # Delete the {%col endfor %} column (shifted by expansion + child changes)
        new_close_col = block.close_col + cols_added + cumulative_child_delta
        self.col_expander.remove_cols(new_close_col, new_close_col)
        self._shift_col_context(new_close_col, -1)

        # Delete the {%col for %} column
        self.col_expander.remove_cols(block.open_col, block.open_col)
        self._shift_col_context(block.open_col, -1)

        return cols_added + cumulative_child_delta - 2

    def _process_col_if_block(
        self, block: ColBlock, context: dict[str, Any]
    ) -> int:
        """Process a column if-block. Returns net column change."""
        condition_expr = block.directive["condition"]

        try:
            expr = self.env.compile_expression(condition_expr)
            result = expr(**context)
        except Exception as e:
            raise TemplateRenderError(
                f"Failed to evaluate condition '{condition_expr}' "
                f"at column {block.open_col}: {e}"
            ) from e

        if result:
            # Condition is true: keep body, remove directive columns.
            # Do NOT render body cells here — they may reference variables
            # from row loops that haven't been expanded yet. Rendering is
            # deferred to the row-block phase or _render_remaining_cells.
            nested_blocks = self._scan_col_blocks_in_range(
                block.body_start, block.body_end
            )
            child_delta = self._process_col_blocks(nested_blocks, context)

            # Remove close column first (higher number), then open column
            adjusted_close = block.close_col + child_delta
            self.col_expander.remove_cols(adjusted_close, adjusted_close)
            self._shift_col_context(adjusted_close, -1)
            self.col_expander.remove_cols(block.open_col, block.open_col)
            self._shift_col_context(block.open_col, -1)
            return child_delta - 2
        else:
            # Condition is false: remove entire block
            delta = self.col_expander.remove_cols(block.open_col, block.close_col)
            self._shift_col_context(block.open_col, delta)
            return delta

    # --- Phase D: Render expression cells ---

    def _render_row_range(
        self, start_row: int, end_row: int, context: dict[str, Any]
    ) -> None:
        """Render all template expression cells in the given row range."""
        for row in range(start_row, end_row + 1):
            for col in range(1, (self.ws.max_column or 1) + 1):
                cell = self.ws.cell(row=row, column=col)
                if not has_template_tag(cell.value):
                    continue
                if is_block_tag(cell.value) or is_col_block_tag(cell.value):
                    continue
                merged = self._merge_col_context(col, context)
                self._render_cell(cell, merged)

    def _render_remaining_cells(self, context: dict[str, Any]) -> None:
        """Render any remaining cells with template tags (plain rows)."""
        for row in range(1, (self.ws.max_row or 0) + 1):
            for col in range(1, (self.ws.max_column or 1) + 1):
                cell = self.ws.cell(row=row, column=col)
                if not has_template_tag(cell.value):
                    continue
                if is_block_tag(cell.value) or is_col_block_tag(cell.value):
                    continue
                merged = self._merge_col_context(col, context)
                self._render_cell(cell, merged)

    def _render_cell(self, cell, context: dict[str, Any]) -> None:
        """Render a single cell's value."""
        value = cell.value
        if not isinstance(value, str):
            return

        # Pure expression: {{ expr }} — preserve Python type
        m = cell_has_only_expression(value)
        if m:
            expr_str = m.group(1)
            try:
                expr = self.env.compile_expression(
                    expr_str, undefined_to_none=False
                )
                result = expr(**context)
                # Force undefined errors for strict mode
                if isinstance(result, Undefined):
                    str(result)
                    result = None
                cell.value = result
            except Exception as e:
                raise TemplateRenderError(
                    f"Failed to render expression '{expr_str}' in cell "
                    f"{cell.coordinate}: {e}"
                ) from e
            return

        # Mixed content: render as string template
        try:
            tpl = self.env.from_string(value)
            rendered = tpl.render(context)
            cell.value = rendered
        except Exception as e:
            raise TemplateRenderError(
                f"Failed to render cell {cell.coordinate}: {e}"
            ) from e
