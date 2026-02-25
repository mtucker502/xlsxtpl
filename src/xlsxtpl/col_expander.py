from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet

from .cell_utils import copy_cell, copy_column_dimensions


class ColExpander:
    """Handles column duplication and deletion for column-level loop/conditional expansion."""

    def __init__(self, ws: Worksheet) -> None:
        self.ws = ws

    def expand_for_loop(
        self, body_start: int, body_end: int, iteration_count: int
    ) -> int:
        """Duplicate the body columns (body_start..body_end) for each iteration.

        The original body columns serve as iteration 0. Additional copies are
        inserted immediately after the original block for iterations 1..N-1.

        Returns the number of net columns added (0 if iteration_count <= 1).
        """
        if iteration_count <= 1:
            return 0

        body_col_count = body_end - body_start + 1
        cols_to_insert = (iteration_count - 1) * body_col_count

        # Insert blank columns right after the original body
        self.ws.insert_cols(body_end + 1, cols_to_insert)

        # Copy template body columns into each new iteration
        max_row = self.ws.max_row or 1
        for iteration in range(1, iteration_count):
            for offset in range(body_col_count):
                src_col = body_start + offset
                tgt_col = body_end + 1 + (iteration - 1) * body_col_count + offset

                copy_column_dimensions(self.ws, src_col, tgt_col)

                for row in range(1, max_row + 1):
                    src_cell = self.ws.cell(row=row, column=src_col)
                    tgt_cell = self.ws.cell(row=row, column=tgt_col)
                    copy_cell(src_cell, tgt_cell)

        return cols_to_insert

    def remove_cols(self, start: int, end: int) -> int:
        """Delete columns from start to end (inclusive). Returns negative offset."""
        count = end - start + 1
        self.ws.delete_cols(start, count)
        return -count
