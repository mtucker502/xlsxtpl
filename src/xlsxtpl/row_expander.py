from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet

from .cell_utils import copy_cell, copy_row_dimensions


class RowExpander:
    """Handles row duplication and deletion for loop/conditional expansion."""

    def __init__(self, ws: Worksheet) -> None:
        self.ws = ws

    def expand_for_loop(
        self, body_start: int, body_end: int, iteration_count: int
    ) -> int:
        """Duplicate the body rows (body_start..body_end) for each iteration.

        The original body rows serve as iteration 0. Additional copies are
        inserted immediately after the original block for iterations 1..N-1.

        Returns the number of net rows added (0 if iteration_count <= 1).
        """
        if iteration_count <= 1:
            return 0

        body_row_count = body_end - body_start + 1
        rows_to_insert = (iteration_count - 1) * body_row_count

        # Insert blank rows right after the original body
        self.ws.insert_rows(body_end + 1, rows_to_insert)

        # Copy template body rows into each new iteration
        max_col = self.ws.max_column or 1
        for iteration in range(1, iteration_count):
            for offset in range(body_row_count):
                src_row = body_start + offset
                tgt_row = body_end + 1 + (iteration - 1) * body_row_count + offset

                copy_row_dimensions(self.ws, src_row, tgt_row)

                for col in range(1, max_col + 1):
                    src_cell = self.ws.cell(row=src_row, column=col)
                    tgt_cell = self.ws.cell(row=tgt_row, column=col)
                    copy_cell(src_cell, tgt_cell)

        return rows_to_insert

    def remove_rows(self, start: int, end: int) -> int:
        """Delete rows from start to end (inclusive). Returns negative offset."""
        count = end - start + 1
        self.ws.delete_rows(start, count)
        return -count
